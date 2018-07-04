import click
import openpyxl
from openpyxl import Workbook
import os
import bs4
from bs4 import BeautifulSoup
import urllib.request




@click.command()
@click.argument('inputurl',nargs=1)
@click.argument('outputxlsx',nargs=1)
def parser(inputurl,outputxlsx):



    try:

        """output data"""
        wb = openpyxl.Workbook()
        sheet = wb.active


        """ opening browser"""
        source = urllib.request.urlopen('https://d1b10bmlvqabco.cloudfront.net/attach/inpg92dp42z2zo/hdff4poirlh7i6/io5hun2sdr21/mock_results.html').read()
        soup=BeautifulSoup(source,'html.parser')

        #print(soup)

        i=soup.find_all('h1')[0]
        sheet.title="Marks"
        print(i.text.strip())

        """copying headings"""
        table=soup.find_all('table')[0]
        tr=table.find_all('tr')
        th=table.find_all('th')
        th=th[1:]
        k=1
        print(th)

        for i in th:
            sheet.cell(row=1, column=k).value = i.text.strip()
            k+=1


        """copying data"""

        table = soup.find_all('table')[0]

        rows=1
        cols=1


        for i in table.find_all('tr'):
            td=i.find_all('td')
            cols=1
            td=td[1:]
            for j in td:
                #print(j.text.strip())
                sheet.cell(row=rows, column=cols).value =j.text.strip()
                cols+=1
            rows+=1

    except openpyxl.utils.exceptions.CellCoordinatesException:
        print("cell cooridinate exception")
        return ;
    except openpyxl.utils.exceptions.IllegalCharacterError:
        print("illegal character exception")
        return ;
    except openpyxl.utils.exceptions.InvalidFileException:
        print("invalid file exception ")
        return ;
    except openpyxl.utils.exceptions.NamedRangeException:
        print("file range exception")
        return ;
    except openpyxl.utils.exceptions.ReadOnlyWorkbookException:
        print("read only workbook exception")
        return ;
    except FileNotFoundError:
        print("file not found exception")
        return ;

    except SyntaxError:
        print("syntax error")
        return;
    except ImportError:
        print("import error")
        return;
    except KeyError:
        print("key error")
        return;
    except AttributeError:
        print("attribute error")
        return;






    wb.save(outputxlsx)


if __name__=='__main__':
    parser()
