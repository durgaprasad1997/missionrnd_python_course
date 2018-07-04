import click
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import openpyxl.utils.exceptions




@click.command()
@click.option('--capitalize/--no--capitalize', default=False)
@click.option('--preservestyles/--no--preservestyles', default=False)
@click.argument('source',nargs=1)
@click.argument('destination',nargs=1)
def fileopen(capitalize,preservestyles,source,destination):

    try:
        wb=load_workbook(filename=source)
        sheet_names=wb.sheetnames

        wb1 = Workbook()
        output_sheet = wb1.active

        for sheet in sheet_names:
            s=wb[sheet]
            output_sheet = wb1.create_sheet()
            output_sheet.title = sheet
            i,j,l=(1,1,0)
            for cell in s:
                l=0
                j=1
                while l<len(cell):
                    if (capitalize):
                        output_sheet.cell(row=i, column=j).value = cell[l].value.capitalize()
                    else:
                        output_sheet.cell(row=i, column=j).value = cell[l].value
                    if (preservestyles):
                        output_sheet.cell(row=i, column=j).style = cell[l].style
                    l+=1
                    j+=1
                i+=1


    except openpyxl.utils.exceptions.CellCoordinatesException:
        print("cell cooridinate exception")
        return ;
    except openpyxl.utils.exceptions.IllegalCharacterError:
        print("illegal character exception")
        return ;
    except openpyxl.utils.exceptions.InvalidFileException:
        print("invalid file exception ")
        return ;
    except openpyxl.utils.exceptions.NamedRangeException:
        print("file range exception")
        return ;
    except openpyxl.utils.exceptions.ReadOnlyWorkbookException:
        print("read only workbook exception")
        return ;
    except FileNotFoundError:
        print("file not found exception")
        return ;

    wb1.save(destination)






if __name__=='__main__':
    fileopen()