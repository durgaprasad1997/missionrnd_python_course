import click
import MySQLdb

from openpyxl import load_workbook




@click.group()
@click.pass_context
def maindb(ctx):
    print("Inside the maindb");



@maindb.command()
@click.pass_context
def createdb(ctx):

    print("creating")

    try:
        db = MySQLdb.connect("localhost", "root", "mysql")
        cursor=db.cursor()
        q="CREATE SCHEMA IF NOT EXISTS summer_pre_assign5_schema"
        cursor.execute(q)


        db = MySQLdb.connect("localhost", "root", "mysql","summer_pre_assign5_schema")
        cursor = db.cursor()



        q="""CREATE TABLE IF NOT EXISTS `summer_pre_assign5_schema`.`current` (
  `name` VARCHAR(45) NULL,
  `college` VARCHAR(45) NULL,
  `emailid` VARCHAR(45) NULL,
  `dbnames` VARCHAR(45) NOT NULL,
  PRIMARY KEY (`dbnames`));"""
        cursor.execute(q)

        q="""CREATE TABLE IF NOT EXISTS `summer_pre_assign5_schema`.`colleges` (
  `collegename` VARCHAR(45) NULL,
  `acronym` VARCHAR(45) NOT NULL,
  `location` VARCHAR(45) NULL,
  `contact` VARCHAR(45) NULL,
  PRIMARY KEY (`acronym`));"""
        cursor.execute(q)

        q="""CREATE TABLE IF NOT EXISTS `summer_pre_assign5_schema`.`marks` (
  `studentname` VARCHAR(45) NOT NULL,
  `college` VARCHAR(45) NOT NULL,
  `transform` INT NULL,
  `fromcustombase26` INT NULL,
  `getpiglatin` INT NULL,
  `topchars` INT NULL,
  `total` INT NULL,
  PRIMARY KEY (`studentname`));"""
        cursor.execute(q)

        q="""ALTER TABLE `summer_pre_assign5_schema`.`marks` 
ADD CONSTRAINT `studentname`
  FOREIGN KEY (`studentname`)
  REFERENCES `summer_pre_assign5_schema`.`current` (`dbnames`)
  ON DELETE NO ACTION
  ON UPDATE NO ACTION;"""
        cursor.execute(q)


        q="""ALTER TABLE `summer_pre_assign5_schema`.`marks` 
ADD INDEX `college_idx` (`college` ASC) VISIBLE;
;
ALTER TABLE `summer_pre_assign5_schema`.`marks` 
ADD CONSTRAINT `college`
  FOREIGN KEY (`college`)
  REFERENCES `summer_pre_assign5_schema`.`colleges` (`acronym`)
  ON DELETE NO ACTION
  ON UPDATE NO ACTION;"""

        cursor.execute(q)

    except MySQLdb.Error:
        print("Error occured while creating schema")
        return ;
    except MySQLdb.DatabaseError:
        print("Data base Error occured while creating schema")
        return ;
    except MySQLdb.IntegrityError:
        print("Integrity Error occured while creating schema")
        return ;
    except MySQLdb.NotSupportedError:
        print("not supported Error occured while creating schema")
        return ;






    print("sucessfully created")




@maindb.command()
@click.pass_context
def dropdb(ctx):
    print("dropping the data base")
    try:
        db = MySQLdb.connect("localhost", "root", "mysql", "summer_pre_assign5_schema")
        cursor = db.cursor()
        q="""DROP DATABASE `summer_pre_assign5_schema`;"""
        cursor.execute(q)

    except MySQLdb.Error:
        print("Error occured while dropping schema")
        return ;
    except MySQLdb.DatabaseError:
        print("Data base Error occured while dropping schema")
        return ;
    except MySQLdb.IntegrityError:
        print("Integrity Error occured while dropping schema")
        return ;
    except MySQLdb.NotSupportedError:
        print("not supported Error occured while dropping schema")
        return ;

    print("successfully dropped")

@maindb.command()
@click.pass_context
def importdata(ctx):
    print("started importing data  please wait.......................")

    try:
        db = MySQLdb.connect("localhost", "root", "mysql", "summer_pre_assign5_schema")
        cursor = db.cursor()


        workbook = load_workbook('C:/PythonCourse/students.xlsx')

        worksheet = workbook.get_sheet_by_name('Current')

        query = "INSERT INTO current (name,college,emailid,dbnames) VALUES (%s,%s,%s,%s)"

        for row in worksheet.iter_rows():
            s1=row[0].value
            s2 = row[1].value
            s3 =row[2].value
            s4 = row[3].value
            s4=s4.lower()

            cursor.execute(query, (s1,s2,s3,s4))
            db.commit()
    except MySQLdb.Error:
        print("Error occured while inserting values")
        return ;
    except MySQLdb.DatabaseError:
        print("Data base Error occured while inserting values")
        return ;
    except MySQLdb.IntegrityError:
        print("Integrity Error occured while inserting values")
        return ;
    except MySQLdb.NotSupportedError:
        print("not supported Error occured while inserting values")
        return ;

    print("current table successfully added")


    try:

        worksheet = workbook.get_sheet_by_name('Colleges')
        query = "INSERT INTO colleges (collegename,acronym,location,contact) VALUES (%s,%s,%s,%s)"

        first=0
        for row in worksheet.iter_rows():
            s1 = row[0].value
            s2 = row[1].value
            s3 = row[2].value
            s4 = row[3].value
        
            if first==0:
                first=1
            else:
                cursor.execute(query, (s1, s2, s3, s4))

        db.commit()
    except MySQLdb.Error:
        print("Error occured while inserting values")
        return;
    except MySQLdb.DatabaseError:
        print("Data base Error occured while inserting values")
        return;
    except MySQLdb.IntegrityError:
        print("Integrity Error occured while inserting values")
        return;
    except MySQLdb.NotSupportedError:
        print("not supported Error occured while inserting values")
        return;

    print("college table successfully added")




    try:

        workbook = load_workbook('C:/PythonCourse/summer_pre_assign3_output.xlsx')
        worksheet = workbook.get_sheet_by_name('Marks')
        query = "INSERT INTO marks (studentname,college,transform,fromcustombase26,getpiglatin,topchars,total) VALUES (%s,%s,%s,%s,%s,%s,%s)"


        i=0
        for row in worksheet.iter_rows():
            #print(row[0].value,row[1].value,row[2].value,row[3].value,row[4].value,row[5].value)


            if(i!=0):
                #print(row)
            
                x=row[0].value
                t=x.split('_')

                s1 = t[2].lower()
                s2 = t[1]
                s3 = row[1].value
                s4 = row[2].value
                s5 = row[3].value
                s6 = row[4].value
                s7 = row[5].value

                q='SELECT COUNT(*) FROM colleges WHERE acronym=%s'
                ans=0
                cursor.execute(q,(s2,))
                ans=cursor.fetchall()
                #print(ans)
                if(ans[0][0]!=0):
                    cursor.execute(query, (s1,s2,s3,s4,s5,s6,s7))

            else:
                i=1

        db.commit()
    except MySQLdb.Error:
        print("Error occured while inserting values")
        return;
    except MySQLdb.DatabaseError:
        print("Data base Error occured while inserting values")
        return;
    except MySQLdb.IntegrityError:
        print("Integrity Error occured while inserting values")
        return;
    except MySQLdb.NotSupportedError:
        print("not supported Error occured while inserting values")
        return;

    print("marks table successfully added")

    print("your data is succesfully imported to database")




@maindb.command()
@click.pass_context
def collegestats(ctx):

    try:
        db = MySQLdb.connect("localhost", "root", "mysql", "summer_pre_assign5_schema")
        cursor = db.cursor()
        query='SELECT college,COUNT(*),MIN(total),MAX(total),AVG(total) FROM marks GROUP BY college'
        cursor.execute(query)

        s="college-acronym".center(40)
        s+="count".center(40)
        s+="min".center(40)
        s+="max".center(40)
        s+='avg'.center(40)
        print(s)
        #print("college-ACRONYM     count    min     max      avg")
        ans=cursor.fetchall()
        for i in ans:
            print(str(i[0]).center(40),str(i[1]).center(40),str(i[2]).center(40),str(i[3]).center(40),str(i[4]).center(40))
    except MySQLdb.Error:
        print("Error occured while analysing data")
        return;
    except MySQLdb.DatabaseError:
        print("Data base Error occured while analysing data")
        return;
    except MySQLdb.IntegrityError:
        print("Integrity Error occured while analysing data")
        return;
    except MySQLdb.NotSupportedError:
        print("not supported Error occured while analysing data")
        return;

    print("stats  successfully printed")





if __name__=='__main__':


    maindb(obj={})

